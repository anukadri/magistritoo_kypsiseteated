""" 4. september 2025
Anu Kadri Uustalu magistritöö programm
aitab vastata küsimustele "Millised on tüüpilised küpsiseteadete omadused?"
selle programmi väljundi abil saab visualiseerida erinevaid küpsiseteadete anatoomiaid

väljund: tekstifail, mis annab andmeid era- ja avaliku sektori erinevate küpsiseteadete omaduste kombinatsioonide kohta"""

from openpyxl.workbook import Workbook as wb
from openpyxl import load_workbook
from pathlib import Path


# avab faili
path = r"Uustalu_MA_andmestik.xlsx" # andmestiku faili path
wbAndmestik = load_workbook(path)
wsAndmestik = wbAndmestik["Andmestik"] #worksheet


# millises veerus on milline info

tervikTekstiLeidja = "A"
veergID = "I"
veergPealkiri = "BT"
veergLylitid = "BU" # sisaldab ka märkeruute
veergNupud = "BV"
veergLingid = "BW"
veergNuppLink = "BX"
veergSektor = "BK"

idKood = None
pealkiri = None
lylitid = None
nupud = None
lingid = None
nuppLink = None
sektor = None


# loeb kokku kui mitu nende omadustega küpsiseteadet andmestikus kokku on
# järjekord: pealkiri, lülitid, nupud, lingid, nupp-lingid
kokkuLoendur = [0, 0, 0, 0, 0]
eraLoendur = [0, 0, 0, 0, 0] 
avalikLoendur = [0, 0, 0, 0, 0]

kokkuOmadused = {}
eraOmadused = {}
avalikOmadused = {}

# leiab, kas antud id-ga pisitekstil on pealkiri, lüliteid, nuppe, linke, nupplinke
# pealkirja lahtris on tabelis "jah" või "ei", teistel arv. Teiste puhul kui on 0, saab vastuseks "ei", muul juhul "jah"

rida = 1
while True:
    rida += 1
    lahter = "".join((tervikTekstiLeidja, str(rida)))
    tervikTekst = ws[lahter].value
    
    if tervikTekst == None:
        break
    if tervikTekst == "na":
        continue
    
    
    idKood = ws["".join((veergID, str(rida)))].value

    
    #print(idKood)
    
    sisuSektor = ws["".join((veergSektor, str(rida)))].value
    if sisuSektor == "era":
        sektor = "era"
    elif sisuSektor == "avalik":
        sektor = "avalik"
    else:
        print(sisuSektor)
    
    sisuPealkiri = ws["".join((veergPealkiri, str(rida)))].value
    if sisuPealkiri == "jah":
        pealkiri = "jah"
        kokkuLoendur[0] += 1
        if sektor == "avalik":
            avalikLoendur[0] += 1
        elif sektor == "era":
            eraLoendur[0] += 1
    else:
        pealkiri = "ei"
        
    sisuLylitid = ws["".join((veergLylitid, str(rida)))].value
    if sisuLylitid == 0:
        lylitid = "ei"
    else:
        lylitid = "jah"
        kokkuLoendur[1] += 1
        if sektor == "avalik":
            avalikLoendur[1] += 1
        elif sektor == "era":
            eraLoendur[1] += 1
        
    sisuNupud = ws["".join((veergNupud, str(rida)))].value
    if sisuNupud == 0:
        nupud = "ei"
    else:
        nupud = "jah"
        kokkuLoendur[2] += 1
        if sektor == "avalik":
            avalikLoendur[2] += 1
        elif sektor == "era":
            eraLoendur[2] += 1
        
    sisuLingid = ws["".join((veergLingid, str(rida)))].value
    if sisuLingid == 0:
        lingid = "ei"
    else:
        lingid = "jah"
        kokkuLoendur[3] += 1
        if sektor == "avalik":
            avalikLoendur[3] += 1
        elif sektor == "era":
            eraLoendur[3] += 1
        
    sisuNuppLink = ws["".join((veergNuppLink, str(rida)))].value
    if sisuNuppLink == 0:
        nuppLink = "ei"
    else:
        nuppLink = "jah"
        kokkuLoendur[4] += 1
        if sektor == "avalik":
            avalikLoendur[4] += 1
        elif sektor == "era":
            eraLoendur[4] += 1
        
    kokkuOmadused[idKood] = pealkiri, lylitid, nupud, lingid, nuppLink
    if sektor == "era":
        eraOmadused[idKood] = pealkiri, lylitid, nupud, lingid, nuppLink
    if sektor == "avalik":
        avalikOmadused[idKood] = pealkiri, lylitid, nupud, lingid, nuppLink

"""
print(f"Kokku on erasektori tekste {len(eraOmadused)}")
print("Siin on erasektori pisitekstid koos omadustega järjekorras pealkiri, lülitid, nupud, lingid, nupplink")
print(eraOmadused)
print("\n")

print(f"Kokku on avaliku sektori tekste {len(avalikOmadused)}")
print("Siin on avaliku sektori pisitekstid koos omadustega järjekorras pealkiri, lülitid, nupud, lingid, nupplink")
print(avalikOmadused)
"""

    
### tekstid organiseeritakse ümber väärtuste, mitte ID kaupa
kokkuOmadusteKaupa = {}

for v6ti, v22rtus in kokkuOmadused.items():
    if not v22rtus in kokkuOmadusteKaupa:
        kokkuOmadusteKaupa[v22rtus] = [v6ti] 
    else:
        kokkuOmadusteKaupa[v22rtus].append(v6ti)

eraOmadusteKaupa = {}

for v6ti, v22rtus in eraOmadused.items():
    if not v22rtus in eraOmadusteKaupa:
        eraOmadusteKaupa[v22rtus] = [v6ti] 
    else:
        eraOmadusteKaupa[v22rtus].append(v6ti)
        
        
avalikOmadusteKaupa = {}

for v6ti, v22rtus in avalikOmadused.items():
    if not v22rtus in avalikOmadusteKaupa:
        avalikOmadusteKaupa[v22rtus] = [v6ti] 
    else:
        avalikOmadusteKaupa[v22rtus].append(v6ti)
        


# kirjutab faili tulemused

fail = open("Kysimus1_omaduste_tulemused.txt", "w")
fail.write("Sõnastiku järjendid on järjekorras: pealkiri, lülitid-märkeruudud, nupud, lingid, nupp-lingi vahepealsed")

### mõlemad sektorid kokku
fail.write("\n\n")
fail.write("MÕLEMAD SEKTORID KOKKU")

fail.write("\n\n")
fail.write(f"Kokku on tekste {len(kokkuOmadused)}")
fail.write("\n\n")
fail.write(f"Erinevaid küpisiseteate malle on {len(kokkuOmadusteKaupa)}")
fail.write("\n")
fail.write("Siin on kõik erinevad küpsiseteadete mallid koos nendesse mallidesse kuuluvad teadete arvuga")

for v6ti, v22rtus in sorted(kokkuOmadusteKaupa.items()):
    fail.write("\n")
    fail.write(str(v6ti))
    fail.write("\t")
    fail.write(str(len(v22rtus)))
         

fail.write("\n\n")
fail.write("Siin on kõik erinevad küpsiseteadete mallid koos pisiteksti id-ga")

for v6ti, v22rtus in kokkuOmadusteKaupa.items():
    fail.write("\n")
    fail.write(str(v6ti))
    fail.write("\t")
    fail.write(str(v22rtus))



### erasektor
fail.write("\n\n")
fail.write("ERASEKTOR")

fail.write("\n\n")
fail.write(f"Kokku on erasektori tekste {len(eraOmadused)}")
fail.write("\n\n")
fail.write(f"Erinevaid erasektori küpisiseteate malle on {len(eraOmadusteKaupa)}")
fail.write("\n")
fail.write("Siin on kõik erinevad erasektori küpsiseteadete mallid koos nendesse mallidesse kuuluvad teadete arvuga")

for v6ti, v22rtus in sorted(eraOmadusteKaupa.items()):
    fail.write("\n")
    fail.write(str(v6ti))
    fail.write("\t")
    fail.write(str(len(v22rtus)))
    fail.write("\t")
    fail.write(f"{str(round(len(v22rtus)*100/len(eraOmadused), 1))} %")
         

fail.write("\n\n")
fail.write("Siin on kõik erinevad erasektori küpsiseteadete mallid koos pisiteksti id-ga")

for v6ti, v22rtus in eraOmadusteKaupa.items():
    fail.write("\n")
    fail.write(str(v6ti))
    fail.write("\t")
    fail.write(str(v22rtus))
    
### avalik sektor
fail.write("\n\n")
fail.write("AVALIK SEKTOR")

fail.write("\n\n")
fail.write(f"Kokku on avaliku sektori tekste {len(avalikOmadused)}")
fail.write("\n\n")
fail.write(f"Erinevaid avaliku sektori küpisiseteate malle on {len(avalikOmadusteKaupa)}")
fail.write("\n")
fail.write("Siin on kõik erinevad avaliku sektori küpsiseteadete mallid koos nendesse mallidesse kuuluvad teadete arvuga")

for v6ti, v22rtus in sorted(avalikOmadusteKaupa.items()):
    fail.write("\n")
    fail.write(str(v6ti))
    fail.write("\t")
    fail.write(str(len(v22rtus)))
    fail.write("\t")
    fail.write(f"{str(round(len(v22rtus)*100/len(avalikOmadused), 1))} %")
         

fail.write("\n\n")
fail.write("Siin on kõik erinevad avaliku sektori küpsiseteadete mallid koos pisiteksti id-ga")

for v6ti, v22rtus in avalikOmadusteKaupa.items():
    fail.write("\n")
    fail.write(str(v6ti))
    fail.write("\t")
    fail.write(str(v22rtus))
    
fail.write("\n\n")
fail.write("Mõlema sektori peale on kokku teateid, kus asub vähemalt üks:\n")
fail.write(f"Pealkiri: {str(kokkuLoendur[0])}\t millest erasektoris {str(eraLoendur[0])}\tavalikus sektoris {str(avalikLoendur[0])}\n")
fail.write(f"Lüliti või märkeruut: {str(kokkuLoendur[1])}\t millest erasektoris {str(eraLoendur[1])}\tavalikus sektoris {str(avalikLoendur[1])}\n")
fail.write(f"Nupp: {str(kokkuLoendur[2])}\t millest erasektoris {str(eraLoendur[2])}\tavalikus sektoris {str(avalikLoendur[2])}\n")
fail.write(f"Link: {str(kokkuLoendur[3])}\t millest erasektoris {str(eraLoendur[3])}\tavalikus sektoris {str(avalikLoendur[3])}\n")
fail.write(f"Nupu ja lingi vahepealne asi: {str(kokkuLoendur[4])}\t millest erasektoris {str(eraLoendur[4])}\tavalikus sektoris {str(avalikLoendur[4])}\n")

    
fail.close()

print("Valmis!")

