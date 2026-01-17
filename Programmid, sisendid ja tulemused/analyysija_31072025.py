"""31. juuli 2025
Anu Kadri Uustalu lõputöö programm, mis loeb .xlsx faili sisse,
loeb sisse kõik algtekstide veergude lahtrid
lemmatiseerib, leiab küsimärkide, punktide, hüümärkide arvu, teksti tähemärkide arvu, tingiva kõneviisi olemasolu ning "sina" kasutuse lemmatiseeritud lauses
kirjutab tulemuse uude .xlsx faili
"""

import timeit
from openpyxl.workbook import Workbook as wb
from openpyxl import load_workbook
from pathlib import Path
import re
import stanza
# stanza.download('et')
nlp = stanza.Pipeline('et')

start = timeit.default_timer()


# funktsioon lemmatiseerimiseks
def lemmatiseerija(sisendLause):
    tekst = nlp(sisendLause)
    lemmaLause = []
    for lause in tekst.sentences:
        for sona in lause.words:
            lemma = sona.lemma
            lemmaLause.append(lemma)
    lemmaLause2 = " ".join(lemmaLause)
            
    return lemmaLause2

# funktsioon tinigva kõneviisi leidmiseks
def tingivK6neviis(sisendLause):
    analyys = nlp(sisendLause)
    for lause in analyys.sentences:
        for sona in lause.words:
            if sona.xpos == "V": 
                omadused = sona.feats
                koneviisArray =  re.findall("Mood=[^|]*", omadused)
                for element in koneviisArray:
                        koneviis = element[5:]
                        if koneviis == "Cnd":
                            return "tingiv"
                        else:
                            continue
            else:
                continue
    return "na"

            

path = r"Uustalu_MA_andmestik.xlsx" # andmestiku faili path


# avab dokumendi
wb = load_workbook(path)

ws = wb["Andmestik"] #worksheet


# loeb sisse A veeru lahtrid, töötleb stanza abil ja kirjutab sama rea B-G lahtritesse tulemuse
veergSisend = "C"
veergV2ljund = "BL"
veergSinaTeie = "BM"
veergPunktid = "BO"
veergHyyum2rk = "BQ"
veergKysim2rk = "BP"
veergTingiv = "BR"
veergPikkus = "BS"
rida = 1
while True:
    rida += 1
    lahterSisend = "".join((veergSisend, str(rida)))
    lahterV2ljund = "".join((veergV2ljund, str(rida)))
    sisu = ws[lahterSisend].value
    
    if sisu == None:
        break
    if sisu == "na":
        toodeldudTekst = "na"
    elif len(sisu) <= 2:
        toodeldudTekst = "liiga lühike tekst!!!"
    else:
        toodeldudTekst = lemmatiseerija(sisu)
        
    ws[lahterV2ljund] = toodeldudTekst 

    # loeb kokku mitu punkti lauses on
    punktiLahter = "".join((veergPunktid, str(rida)))
    punkte = sisu.count(".")
    if punkte == 0:
        ws[punktiLahter] = "na"
    else:
        ws[punktiLahter] = punkte

    # loeb kokku mitu hüüumärki lauses on
    hyyuLahter = "".join((veergHyyum2rk, str(rida)))
    hyyu = sisu.count("!")
    if hyyu == 0:
        ws[hyyuLahter] = "na"
    else:
        ws[hyyuLahter] = hyyu


    # loeb kokku mitu küsimärki lauses on
    kysiLahter = "".join((veergKysim2rk, str(rida)))
    kysi = sisu.count("?")
    if kysi == 0:
        ws[kysiLahter] = "na"
    else:
        ws[kysiLahter] = kysi


    # leiab kas lemmatiseeritud tekstis on "sina" või "teie" ja kirjutab vastuse C veergu
    sinaTeieLahter = "".join((veergSinaTeie, str(rida)))

    if " sina" in toodeldudTekst:
        ws[sinaTeieLahter] = "sina"
    elif "sina " in toodeldudTekst:
        ws[sinaTeieLahter] = "sina"
    elif "teie " in toodeldudTekst:
        ws[sinaTeieLahter] = "teie"
    elif " teie" in toodeldudTekst:
        ws[sinaTeieLahter] = "teie"
    else:
        ws[sinaTeieLahter] = "na"

    #print(sisu)
    #print(toodeldudTekst)

    # leiab kas lauses on tingivat kõneviisi
    tingivLahter = "".join((veergTingiv, str(rida)))
    ws[tingivLahter] = tingivK6neviis(sisu)

    # lisab teksti pikkuse
    pikkusLahter = "".join((veergPikkus, str(rida)))
    ws[pikkusLahter] = len(sisu)


    # teeb sama üldise teksti kohta
veergSisend = "B"
veergSinaTeie = "BM"
veergPunktid = "BO"
veergHyyum2rk = "BQ"
veergKysim2rk = "BP"
veergTingiv = "BR"
veergPikkus = "BS"
rida = 1
while True:
    rida += 1
    lahterSisend = "".join((veergSisend, str(rida)))
    sisu = ws[lahterSisend].value

    if sisu == None:
        break
    if sisu == "na":
        continue

    toodeldudTekst = lemmatiseerija(sisu)
    # leiab kas lemmatiseeritud tekstis on "sina" või "teie" ja kirjutab vastuse C veergu
    sinaTeieLahter = "".join((veergSinaTeie, str(rida)))

    if " sina" in toodeldudTekst:
        ws[sinaTeieLahter] = "sina"
    elif "sina " in toodeldudTekst:
        ws[sinaTeieLahter] = "sina"
    elif "teie " in toodeldudTekst:
        ws[sinaTeieLahter] = "teie"
    elif " teie" in toodeldudTekst:
        ws[sinaTeieLahter] = "teie"
    else:
        ws[sinaTeieLahter] = "na"

    # loeb kokku mitu punkti lauses on
    punktiLahter = "".join((veergPunktid, str(rida)))
    punkte = sisu.count(".")
    if punkte == 0:
        ws[punktiLahter] = "na"
    else:
        ws[punktiLahter] = punkte

    # loeb kokku mitu hüüumärki lauses on
    hyyuLahter = "".join((veergHyyum2rk, str(rida)))
    hyyu = sisu.count("!")
    if hyyu == 0:
        ws[hyyuLahter] = "na"
    else:
        ws[hyyuLahter] = hyyu


    # loeb kokku mitu küsimärki lauses on
    kysiLahter = "".join((veergKysim2rk, str(rida)))
    kysi = sisu.count("?")
    if kysi == 0:
        ws[kysiLahter] = "na"
    else:
        ws[kysiLahter] = kysi

    # leiab kas lauses on tingivat kõneviisi
    tingivLahter = "".join((veergTingiv, str(rida)))
    ws[tingivLahter] = tingivK6neviis(sisu)

    # lisab teksti pikkuse
    pikkusLahter = "".join((veergPikkus, str(rida)))
    ws[pikkusLahter] = len(sisu)


 


# salvestab andmed uude faili
pathUusFail = r"MA_andmestik_analyysitud.xlsx" #teeb uue nimega uue faili kuhu sisse on kirjutatud vajalikud asjad
wb.save(pathUusFail)   
print("Valmis!")

stop = timeit.default_timer()
print('Time: ', stop - start)
