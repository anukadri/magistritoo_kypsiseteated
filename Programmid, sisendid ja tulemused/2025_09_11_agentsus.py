""" 11. september 2025
Anu Kadri Uustalu magistritöö programm
aitab vastata küsimustele "Kas esimesel kihil saab keelduda? Kas esimesel kihil saab ise valida?"
selle programmi väljundi abil saab visualiseerida Sankey diagrammi abil küpsiseteadete nuppude tegevusi

Programm loeb sisse MA töö andmestiku, tegeleb lisamärgendustega "Nuppe ja nupp-linke kokku",
                                                                "(Nupud ja nupp-lingid) on visuaalselt võrdsed
                                                                "Esimesel kihil keeldumise nupp/nupplink"
                                                                "Esimesel kihil nupu kaudu valmise võimalus (lülitid või nupp "vali ise")"
                                                                ("Mis tegevus on esilduval (või ainsal) nupul?")


väljund: .txt fail, tabuleeritud andmetega, mis on vajaliku struktuuriga, et sellest saaks sankey diagrammi teha"""


""" Sisendiks tuleb andmestiku fail ning ka fail 2025_09_11_tyhi_tabel_Sankey.xlsx, mis sisaldab vahelehel "Rajad" radu, mida pidi Sankey diagrammi teha
Lehel "Tabel" on ilma väärtusteta tabel ning selle programmi abil kirjutatakse see tabel koos väärtustega uude faili
Kui see programm on oma ülesehituselt segane, soovitan vaadata faili 2025_09_11_tyhi_tabel_Sankey.xlsx"""

from openpyxl.workbook import Workbook as wb
from openpyxl import load_workbook
from pathlib import Path


def v22rtuseKirjutaja(sektor, lahter):
    v22rtus[lahter-2] += 1
    if sektor == "avalik":
        avalikV22rtus[lahter-2] += 1
    elif sektor == "era":
        eraV22rtus[lahter-2] += 1
    return None

# avab faili
path = r"Uustalu_MA_andmestik.xlsx" # andmestiku faili path
wbAndmestik = load_workbook(path)
wsAndmestik = wbAndmestik["Andmestik"] #worksheet

path = r"2025_09_11_tyhi_tabel_Sankey.xlsx" #lisafaili path
wbTabel = load_workbook(path)
wsTabel = wbTabel["Tabel"] #worksheet


# millises veerus on milline info

tervikTekstiLeidja = "A"
veergID = "I"
veergSektor = "BK"
veergNuppArv = "BY"
veergVisuaalseltVordsed = "BZ"
veergKeeldumine = "CA"
veergValimine = "CB"
#veergEsilduvNupp = "CC"

# tühjad muutujad lõpliku tabeli lahtrite nimedega, mille kogus hakkab muutuma, kui vastavaid instantse lisandub

"""v22rtus = [c2, c3, c4, c5, c6, c7,
           c8, c9, c10, c11, c12, c13, c14,
           c15, c16, c17, c18, c19]"""
v22rtus = [0, 0, 0, 0, 0, 0,
           0, 0, 0, 0, 0, 0, 0,
           0, 0, 0, 0, 0]
eraV22rtus = [0, 0, 0, 0, 0, 0,
           0, 0, 0, 0, 0, 0, 0,
           0, 0, 0, 0, 0]
avalikV22rtus = [0, 0, 0, 0, 0, 0,
           0, 0, 0, 0, 0, 0, 0,
           0, 0, 0, 0, 0]



# leiab kõik terviktekstid ning nende ID koodid

rida = 1
while True:
    rida += 1
    lahter = "".join((tervikTekstiLeidja, str(rida))) # leiab lahtri, kust teksti leida, nt A1
    tervikTekst = wsAndmestik[lahter].value
    
    if tervikTekst == None: # kood lõppeb ära, kui tuleb ette tühi lahter
        break
    if tervikTekst == "na": # programm jätab vahele kõik "na" read ehk lause, mitte tervikteksti read
        continue
    
    
    idKood = wsAndmestik["".join((veergID, str(rida)))].value
    
    #print(idKood)
    
    sisuSektor = wsAndmestik["".join((veergSektor, str(rida)))].value
    if sisuSektor == "era":
        sektor = "era"
    elif sisuSektor == "avalik":
        sektor = "avalik"
    else:
        print(sisuSektor)
    
    nuppArv = wsAndmestik["".join((veergNuppArv, str(rida)))].value
    
    valimine = wsAndmestik["".join((veergValimine, str(rida)))].value
    visuaalseltVordsed = wsAndmestik["".join((veergVisuaalseltVordsed, str(rida)))].value
    keeldumine = wsAndmestik["".join((veergKeeldumine, str(rida)))].value
    #print(nuppArv)
    
    ### NUPPE EI OLE
    if nuppArv == 0:
        #c2 += 1
        v22rtuseKirjutaja(sektor, 2)
    ### ÜKS NUPP
    elif nuppArv == 1:
        if valimine != "lüliti või märkeruut":
            #c3 += 1
            #c5 += 1
            v22rtuseKirjutaja(sektor, 3)
            v22rtuseKirjutaja(sektor, 5)
        elif valimine == "lüliti või märkeruut":
            #c4 += 1
            #c6 += 1
            v22rtuseKirjutaja(sektor, 4)
            v22rtuseKirjutaja(sektor, 6)
        else:
            print(idKood)
    ### ROHKEM KUI ÜKS NUPP
    elif nuppArv > 1:
        if visuaalseltVordsed == "jah":
            #c7 += 1
            v22rtuseKirjutaja(sektor, 7)
            
            # teatel on lülitid-märkeruudud
            if valimine == "lüliti või märkeruut": 
                #c6 += 1
                #c9 += 1
                v22rtuseKirjutaja(sektor, 6)
                v22rtuseKirjutaja(sektor, 9)
                
            #ei saa keelduda EGA valida
            elif keeldumine == "ei" and valimine == "ei":
                #c10 += 1
                #c13 += 1
                v22rtuseKirjutaja(sektor, 10)
                v22rtuseKirjutaja(sektor, 13)
            
            #saab keelduda JA valida
            elif (keeldumine == "keelduma" or keeldumine == "nõustuma hädavajalikega") and (valimine == "jah (verbid muuda, kohanda, halda, vali jms)" or valimine == "jah, tekst seaded, sätted vms"):
                #c11 += 1
                #c14 += 1
                v22rtuseKirjutaja(sektor, 11)
                v22rtuseKirjutaja(sektor, 14)
                
            # saab ainult keelduda või ainult valida
            else:
                #c12 += 1
                #c15 += 1
                v22rtuseKirjutaja(sektor, 12)
                v22rtuseKirjutaja(sektor, 15)
                
        elif visuaalseltVordsed == "ei":
            #c8 += 1
            v22rtuseKirjutaja(sektor, 8)
            if keeldumine == "ei" and valimine == "ei":
                #c16 += 1
                #c18 += 1
                v22rtuseKirjutaja(sektor, 16)
                v22rtuseKirjutaja(sektor, 18)
            else:
                #c17 += 1
                #c19 += 1
                v22rtuseKirjutaja(sektor, 17)
                v22rtuseKirjutaja(sektor, 19)
        else:
            print("Probleem teatega:")
            print(idKood)
            
            
# kirjutab faili tulemused

fail = open("Kysimus9_Sankey_tulemused.txt", "w")

fail.write("Mõlemad sektorid kokku\nFrom\tTo\tKogus\n")

lahter1 = ["A2", "A3", "A4", "A5", "A6", "A7",
           "A8", "A9", "A10", "A11", "A12",
           "A13", "A14", "A15", "A16", "A17", "A18",
           "A19", "A20", "A21", "A22", "A23", "A24", "A25", "A26"]
lahter2 = ["B2", "B3", "B4", "B5", "B6", "B7",
           "B8", "B9", "B10", "B11", "B12",
           "B13", "B14", "B15", "B16", "B17", "B18",
           "B19", "B20", "B21", "B22", "B23", "B24", "B25", "B26"]


i = 0
for i in range(18):
    fail.write(f"{wsTabel[lahter1[i]].value}\t{wsTabel[lahter2[i]].value}\t{v22rtus[i]}\n")
    i += 1

fail.write("\n\nAvalik sektor\nFrom\tTo\tKogus\n")

i = 0
for i in range(18):
    fail.write(f"{wsTabel[lahter1[i]].value}\t{wsTabel[lahter2[i]].value}\t{avalikV22rtus[i]}\n")
    i += 1
    
fail.write("\n\nErasektor\nFrom\tTo\tKogus\n")

i = 0
for i in range(18):
    fail.write(f"{wsTabel[lahter1[i]].value}\t{wsTabel[lahter2[i]].value}\t{eraV22rtus[i]}\n")
    i += 1

fail.close()

print("Valmis!")



        

        
        
        



