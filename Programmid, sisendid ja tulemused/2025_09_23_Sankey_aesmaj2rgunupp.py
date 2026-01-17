""" 23. september 2025
Anu Kadri Uustalu magistritöö programm
aitab vastata küsimustele "Millised on esmajärgu nuppude tekstid?"
selle programmi väljundi abil saab visualiseerida Sankey diagrammi abil küpsiseteadete nuppude tegevusi

Programm loeb sisse MA töö andmestiku, tegeleb lisamärgendustega "Mis tegevus on esilduval (või ainsal) nupul?"
                                                                "Mitu esilduvat nuppu?"


väljund: .txt fail, tabuleeritud andmetega, mis on vajaliku struktuuriga, et sellest saaks sankey diagrammi teha"""


""" Sisendiks tuleb andmestiku fail ning ka fail 2025_09_11_tyhi_tabel_Sankey.xlsx, mis sisaldab vahelehel "Rajad" radu, mida pidi Sankey diagrammi teha
Lehel "Tabel" on ilma väärtusteta tabel ning selle programmi abil kirjutatakse see tabel koos väärtustega uude faili
Kui see programm on oma ülesehituselt segane, soovitan vaadata faili 2025_09_11_tyhi_tabel_Sankey.xlsx"""

from openpyxl.workbook import Workbook as wb
from openpyxl import load_workbook
from pathlib import Path


def v22rtuseKirjutaja(sektor, lahter):
    v22rtus[lahter-2] += 1
    if sektor == "avalik":
        avalikV22rtus[lahter-2] += 1
    elif sektor == "era":
        eraV22rtus[lahter-2] += 1
    return None


# avab faili
path = r"Uustalu_MA_andmestik.xlsx" # andmestiku faili path
wbAndmestik = load_workbook(path)
wsAndmestik = wbAndmestik["Andmestik"] #worksheet

path = r"2025_09_11_tyhi_tabel_Sankey.xlsx" #lisafaili path
wbTabel = load_workbook(path)
wsTabel = wbTabel["Tabel2_esmaj2rk_tekst"] #worksheet


# millises veerus on milline info

tervikTekstiLeidja = "A"
veergID = "I"
veergSektor = "BK"
veergVisuaalseltVordsed = "BZ"
veergLylitid = "BU"
veergEsilduvateArv = "CD"
veergEsilduvTegevus = "CE"


v22rtus = [0, 0, 0, 0, 0, 0, 0, 0,
           0, 0, 0, 0, 0, 0, 0]
eraV22rtus = [0, 0, 0, 0, 0, 0, 0, 0,
           0, 0, 0, 0, 0, 0, 0]
avalikV22rtus = [0, 0, 0, 0, 0, 0, 0, 0,
           0, 0, 0, 0, 0, 0, 0]


# leiab kõik terviktekstid ning nende ID koodid

rida = 1
while True:
    rida += 1
    lahter = "".join((tervikTekstiLeidja, str(rida))) # leiab lahtri, kust teksti leida, nt A1
    tervikTekst = wsAndmestik[lahter].value
    
    if tervikTekst == None: # kood lõppeb ära, kui tuleb ette tühi lahter
        break
    if tervikTekst == "na": # programm jätab vahele kõik "na" read ehk lause, mitte tervikteksti read
        continue
    
    
    idKood = wsAndmestik["".join((veergID, str(rida)))].value
    
    #print(idKood)
    
    sisuSektor = wsAndmestik["".join((veergSektor, str(rida)))].value
    if sisuSektor == "era":
        sektor = "era"
    elif sisuSektor == "avalik":
        sektor = "avalik"
    else:
        print(sisuSektor)
        
        
    visuaalseltVordsed = wsAndmestik["".join((veergVisuaalseltVordsed, str(rida)))].value
    lylitid = wsAndmestik["".join((veergLylitid, str(rida)))].value
    esilduvateArv = wsAndmestik["".join((veergEsilduvateArv, str(rida)))].value
    esilduvTegevus = wsAndmestik["".join((veergEsilduvTegevus, str(rida)))].value

    
    if visuaalseltVordsed != "jah":
        try:
            ### ÜKS ESILDUV NUPP
            if esilduvateArv == 1:
                # lüliteid pole
                if lylitid == 0:
                    v22rtuseKirjutaja(sektor, 3)
                    if esilduvTegevus == "selge-aktsepteeri":
                        v22rtuseKirjutaja(sektor, 6)
                    elif esilduvTegevus == "keelduma":
                        v22rtuseKirjutaja(sektor, 7)
                    elif esilduvTegevus == "muu":
                        v22rtuseKirjutaja(sektor, 8)
                    elif esilduvTegevus == "nõustuma":
                        v22rtuseKirjutaja(sektor, 9)
                    elif esilduvTegevus == "nõustuma kõigiga":
                        v22rtuseKirjutaja(sektor, 10)
                # lülitid on
                else:
                    v22rtuseKirjutaja(sektor, 2)
                    if esilduvTegevus == "nõustuma":
                        v22rtuseKirjutaja(sektor, 4)
                    elif esilduvTegevus == "salvesta (valikud)":
                        v22rtuseKirjutaja(sektor, 5)
                        
            ### MITU ESILDUVAT NUPPU
            elif esilduvateArv == 2 or esilduvateArv == 3:
                # lüliteid pole
                if lylitid == 0:
                    v22rtuseKirjutaja(sektor, 12)
                    if esilduvTegevus == "keelduma ja ok":
                        v22rtuseKirjutaja(sektor, 15)
                    elif esilduvTegevus == "nõustuma ja keelduma":
                        v22rtuseKirjutaja(sektor, 16)

                # lülitid on
                else:
                    v22rtuseKirjutaja(sektor, 11)
                    if esilduvTegevus == "nõustu, keeldu ja nõustu valikutega":
                        v22rtuseKirjutaja(sektor, 13)
                    elif esilduvTegevus == "nõustuma ja keelduma":
                        v22rtuseKirjutaja(sektor, 14)
        except:
            continue
                
                
# kirjutab faili tulemused

fail = open("Kysimus9_Sankey_esmanupp_tulemused.txt", "w")

fail.write("Mõlemad sektorid kokku\nFrom\tTo\tKogus\n")

lahter1 = ["A2", "A3", "A4", "A5", "A6", "A7",
           "A8", "A9", "A10", "A11", "A12",
           "A13", "A14", "A15", "A16"]
lahter2 = ["B2", "B3", "B4", "B5", "B6", "B7",
           "B8", "B9", "B10", "B11", "B12",
           "B13", "B14", "B15", "B16"]


i = 0
for i in range(15):
    fail.write(f"{wsTabel[lahter1[i]].value}\t{wsTabel[lahter2[i]].value}\t{v22rtus[i]}\n")
    i += 1

fail.write("\n\nAvalik sektor\nFrom\tTo\tKogus\n")

i = 0
for i in range(15):
    fail.write(f"{wsTabel[lahter1[i]].value}\t{wsTabel[lahter2[i]].value}\t{avalikV22rtus[i]}\n")
    i += 1
    
fail.write("\n\nErasektor\nFrom\tTo\tKogus\n")

i = 0
for i in range(15):
    fail.write(f"{wsTabel[lahter1[i]].value}\t{wsTabel[lahter2[i]].value}\t{eraV22rtus[i]}\n")
    i += 1

fail.close()

print("Valmis!")
                
    

